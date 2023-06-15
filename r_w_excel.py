#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @Time      :2023/3/13 11:25
# @Author    :Joy
# @FileName  :r_w_excel.py
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
import threading
import os

lock = threading.Lock()
count = 0
max_size = 10  # 单文件最大占用内存


def transpose(matrix):
    # workbook = Workbook()
    # # 获取默认工作表
    # worksheet = workbook.active
    # for i in matrix:
    #     if len(i[4])>32:
    #          worksheet.append(i)
    # workbook.save('32.xlsx')

    # 获取矩阵行数和列数
    rows = len(matrix)
    cols = len(matrix[0])
    # 创建一个新的二维列表用于存储转置后的矩阵
    transposed = [[0 for j in range(rows)] for i in range(cols)]
    # 遍历原矩阵，将每个元素移到对应位置上
    for i in range(rows):
        for j in range(cols):
            transposed[j][i] = matrix[i][j]
    for k in transposed[4:]:
        print(f'{k[0]}: {set(k[1:])}')


def get_filename():
    path = "olddata"

    # 获取目录下所有文件列表
    files = os.listdir(path)
    if files:
        # 打印文件列表
        for file in files:
            if file.split('.')[-1] in ['xlsx', 'xls']:
                return file
    raise FileNotFoundError('-----------找不到文件-----------')


def read_template():
    # 打开Excel文件  获取母版excel数据
    workbook = load_workbook(filename=r'template/template.xlsx')

    # 选择一个工作表
    worksheet = workbook.worksheets[0]

    # 遍历每一行并输出
    template = []
    for row in worksheet.iter_rows(values_only=True):
        template.append(list(row))
    return template


def read_excel():
    filename = get_filename()
    # 打开Excel文件  获取母版excel数据
    workbook = load_workbook(filename=fr'olddata/{filename}')

    # 选择一个工作表
    worksheet = workbook.worksheets[0]

    # 遍历每一行并输出
    excel_data = []
    for row in worksheet.iter_rows(values_only=True):
        excel_data.append(list(row))
    return excel_data


def offset_img(img, col, row):
    """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
    row column 的索引都是从0开始的，我这里要把图片插入到单元格B10
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col, colOff=60000, row=row, rowOff=60000)
    img.anchor = OneCellAnchor(_from=marker, ext=size)


def write_excel(data, i, distance):
    # 创建Workbook对象
    workbook = Workbook()

    # 获取默认工作表
    worksheet = workbook.active

    # 添加模板头部
    for template in read_template():
        worksheet.append(template)

    # 将嵌套列表写入工作表
    if type(distance) is list:
        begin, end = distance
    else:
        begin, end = (i * distance, (i + 1) * distance)
    for index, row in enumerate(data[begin:end]):
        urls = row[0]

        # 将图片外的内容写入excel
        row[0] = ''
        # if len(row[4]) <= 32:
        worksheet.append(row)
        try:
            photos = urls.split(',')
            for ii, url in enumerate(photos):
                _filename = '%s.%s' % (f'p_{ii}_{row[2]}', url.split('/')[-1].split('.')[-1])
                # 获取图片并调整图片大小和单元格大小
                img = Image(rf'img/{_filename}')
                img.width, img.height = (img.width * 72 / img.height, 72)
                offset_img(img, col=0, row=index + 2)

                worksheet.column_dimensions['A'].width = 10 + 2

                worksheet.row_dimensions[index + 3].height = 60 + 12
                worksheet.add_image(img)  # 插入图片
        except FileNotFoundError:
            row[0] = urls
            print(f'找不到图片信息，表格内容:{row}')
        except AttributeError:
            pass

    # 保存Excel文件
    lock.acquire()
    global count
    count += 1
    print(f'当前已完成excel进度：{count}')
    lock.release()
    workbook.save(filename=fr'newdata/example{begin}-{end}.xlsx')


def write_excel_for_size(data):
    current_size = 0
    b_to_e = [0, 0]
    b_to_e_lists = []
    for index, row in enumerate(data):
        if b_to_e[1] > 0:
            b_to_e = [index - 1, 0]
        urls = row[0]
        try:
            photos = urls.split(',')
            photos_size = 0
            for ii, url in enumerate(photos):

                _filename = '%s.%s' % (f'p_{ii}_{row[2]}', url.split('/')[-1].split('.')[-1])
                # 获取图片大小
                size = os.path.getsize(fr'img/{_filename}') / 1024 / 1024
                photos_size += size
            current_size += photos_size
            if index - b_to_e[0] >= 100 or current_size > max_size:  # 不大于100款 且 不大于指定大小
                # print(max_size)
                current_size = photos_size
                b_to_e[1] = index
                b_to_e_lists.append(b_to_e)
        except:
            pass
    b_to_e[1] = len(data)
    b_to_e_lists.append(b_to_e)
    # print(b_to_e_lists)
    for b2e in b_to_e_lists:
        write_excel(data=data, i=0, distance=b2e)


def get_pic(data, i, distance):
    begin, end = (i * distance, (i + 1) * distance)
    for row in data[begin:end]:
        urls = row[0]
        try:
            # _filename = url.split('/')[-1]
            photos = urls.split(',')
            for ii, url in enumerate(photos):
                _filename = '%s.%s' % (f'p_{ii}_{row[2]}', url.split('/')[-1].split('.')[-1])
                if not os.path.exists(fr'img/{_filename}'):
                    res = requests.get(url).content
                    # 保存图片
                    with open(fr'img/{_filename}', 'wb') as f:
                        f.write(res)
        except:
            print(f'找不到图片信息，表格内容:{row}')
        lock.acquire()
        global count
        count += 1
        print(f'当前下载图片进度：{count}/{len(data)}')
        lock.release()


def run(fun, data, distance):
    th_list = []
    for i, j in enumerate(data[::distance]):
        t = threading.Thread(target=fun, args=(data, i, distance))
        th_list.append(t)

    for th_start in th_list:
        th_start.start()

    for th_join in th_list:
        th_join.join()


if __name__ == '__main__':
    print('-----------只读取获取到的第一个excel文件-----------')
    print('-----------开始读取excel数据-----------')
    _data = read_excel()
    # transpose(_data)
    print('-----------读取excel数据完成-----------')
    while True:
        print('请输入:')
        print('1:将图片下载到本地')
        print('2:生成新的款型库导入模板')
        print(f'3:生成新的且文件大小不超过指定大小(默认10M)且不大于100款的款型库导入模板')
        print('4:退出')

        try:
            model = int(input('>>'))
        except ValueError:
            print('输入不合法，请重新输入!')
            continue
        else:
            if model == 1:
                distance = int(input('请输入单个线程下载图片的数量:'))
                run(fun=get_pic, data=_data, distance=distance)  # 获取图片
                break
            elif model == 2:
                distance = int(input('请输入单个文件包含的款式数量:'))
                run(fun=write_excel, data=_data, distance=distance)  # 写入excel
                break
            elif model == 3:
                max_size = int(input('请输入单个文件大小，单位M:'))
                write_excel_for_size(_data)
                break
            elif model == 4:
                break
            else:
                print('输入不合法，请重新输入!')
