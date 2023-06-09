#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @Time      :2023/3/13 11:25
# @Author    :Joy
# @FileName  :r_w_excel.py
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import threading
import os

from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU

lock = threading.Lock()
count = 0


def transpose(matrix):
    # 获取矩阵行数和列数
    rows = len(matrix)
    cols = len(matrix[0])
    # 创建一个新的二维列表用于存储转置后的矩阵
    transposed = [[0 for j in range(rows)] for i in range(cols)]
    # 遍历原矩阵，将每个元素移到对应位置上
    for i in range(rows):
        for j in range(cols):
            transposed[j][i] = matrix[i][j]
    for k in transposed[4:]:
        print(f'{k[0]}: {set(k[1:])}')


def get_filename():
    path = "olddata"

    # 获取目录下所有文件列表
    files = os.listdir(path)
    if files:
        # 打印文件列表
        for file in files:
            if file.split('.')[-1] in ['xlsx', 'xls']:
                return file
    raise FileNotFoundError('-----------找不到文件-----------')


def read_template():
    # 打开Excel文件  获取母版excel数据
    workbook = load_workbook(filename=r'template/template.xlsx')

    # 选择一个工作表
    worksheet = workbook.worksheets[0]

    # 遍历每一行并输出
    template = []
    for row in worksheet.iter_rows(values_only=True):
        template.append(list(row))
    return template


def read_excel():
    filename = get_filename()
    # 打开Excel文件  获取母版excel数据
    workbook = load_workbook(filename=fr'olddata/{filename}')

    # 选择一个工作表
    worksheet = workbook.worksheets[0]

    # 遍历每一行并输出
    excel_data = []
    for row in worksheet.iter_rows(values_only=True):
        excel_data.append(list(row))
    return excel_data


def write_excel(data, i, distance):
    # 创建Workbook对象
    workbook = Workbook()

    # 获取默认工作表
    worksheet = workbook.active

    # 添加模板头部
    for template in read_template():
        worksheet.append(template)

    # 将嵌套列表写入工作表
    begin, end = (i * distance, (i + 1) * distance)
    for index, row in enumerate(data[begin:end]):
        url = row[0]

        # 将图片外的内容写入excel
        row[0] = ''
        worksheet.append(row)
        try:
            # _filename = url.split('/')[-1]
            _filename = '%s.%s' % (row[2], url.split('/')[-1].split('.')[-1])
            # 获取图片并调整图片大小和单元格大小
            img = Image(rf'img/{_filename}')
            # 调整图片大小
            img.width, img.height = (img.width / 4, img.height / 4)
            # 修改单元格宽度和高度
            worksheet.column_dimensions['A'].width = img.width / 5
            worksheet.row_dimensions[index + 3].height = img.height

            # 添加图片并居中显示
            c2e = cm_to_EMU
            p2e = pixels_to_EMU
            size = XDRPositiveSize2D(p2e(img.height), p2e(img.width))
            marker = AnchorMarker(col=0, colOff=10, row=index + 2, rowOff=10)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img, anchor='C10')
        except FileNotFoundError:
            row[0] = url
            print(f'找不到图片信息，表格内容:{row}')
        except AttributeError:
            pass

    # 保存Excel文件
    lock.acquire()
    global count
    count += 1
    print(f'当前已完成excel进度：{count}')
    lock.release()
    workbook.save(filename=f'newdata/example{begin}-{end}.xlsx')


def get_pic(data, i, distance):
    begin, end = (i * distance, (i + 1) * distance)
    for row in data[begin:end]:
        url = row[0]
        try:
            _filename = url.split('/')[-1]
            _filename = '%s.%s' % (row[2], url.split('/')[-1].split('.')[-1])
            res = requests.get(url).content

            # 保存图片
            with open(fr'img/{_filename}', 'wb') as f:
                f.write(res)
        except:
            print(f'找不到图片信息，表格内容:{row}')
        lock.acquire()
        global count
        count += 1
        print(f'当前下载图片进度：{count}/{len(data)}')
        lock.release()


def run(fun, data, distance):
    th_list = []
    for i, j in enumerate(data[::distance]):
        t = threading.Thread(target=fun, args=(data, i, distance))
        th_list.append(t)

    for th_start in th_list:
        th_start.start()

    for th_join in th_list:
        th_join.join()


if __name__ == '__main__':
    print('-----------只读取获取到的第一个excel文件-----------')
    print('-----------开始读取excel数据-----------')
    _data = read_excel()
    # transpose(_data)
    print('-----------读取excel数据完成-----------')
    while True:
        print('请输入:')
        print('1:将图片下载到本地')
        print('2:生成新的款型库导入模板')
        print('3:退出')

        model = int(input('>>'))
        if model == 1:
            while True:
                count = 0
                distance = int(input('请输入单个线程下载图片的数量:'))
                if type(distance) == int:
                    run(fun=get_pic, data=_data, distance=distance)  # 获取图片
                    break
                else:
                    print('输入不合法，请重新输入')
        elif model == 2:
            while True:
                count = 0
                distance = int(input('请输入单个文件包含的款式数量:'))
                if type(distance) == int:
                    run(fun=write_excel, data=_data, distance=distance)  # 写入excel
                    break
                else:
                    print('输入不合法，请重新输入')
        elif model == 3:
            break
        else:
            print('输入不合法，请重新输入')

    # run(fun=get_pic, data=_data, distance=1000)  # 获取图片  distance为单个线程下载图片的数量
    # run(fun=write_excel, data=_data, distance=100)  # 写入excel  distance单个文件包含的款式数量
